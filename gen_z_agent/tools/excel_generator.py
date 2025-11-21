"""
Excel Report Generator
실제 Excel 파일을 생성하는 보고서 생성기
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """
    Generate professional Excel reports with multiple sheets and formatting.
    """

    @staticmethod
    def create_workbook():
        """Create a new Excel workbook."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        return openpyxl.Workbook()

    @staticmethod
    def style_header(worksheet, row: int = 1):
        """
        Apply header styling to a row.

        Args:
            worksheet: Excel worksheet
            row: Row number to style
        """
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

    @staticmethod
    def auto_size_columns(worksheet, min_width: int = 10, max_width: int = 50):
        """
        Auto-size columns based on content.

        Args:
            worksheet: Excel worksheet
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def create_summary_sheet(workbook, analysis_data: Dict[str, Any]) -> None:
        """
        Create a summary sheet with key metrics.

        Args:
            workbook: Excel workbook
            analysis_data: Analysis results
        """
        ws = workbook.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "선거 데이터 분석 요약"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # Metadata
        row = 3
        ws[f'A{row}'] = "분석 ID:"
        ws[f'B{row}'] = analysis_data.get("analysis_id", "N/A")
        row += 1
        ws[f'A{row}'] = "분석 일시:"
        ws[f'B{row}'] = analysis_data.get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row += 1
        ws[f'A{row}'] = "문서명:"
        ws[f'B{row}'] = analysis_data.get("document_name", "N/A")
        row += 2

        # Key metrics
        ws[f'A{row}'] = "핵심 메트릭"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        metrics = analysis_data.get("key_metrics", {})
        for key, value in metrics.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1

        # Style
        ExcelReportGenerator.auto_size_columns(ws)

    @staticmethod
    def create_data_sheet(workbook, sheet_name: str, data: List[Dict], headers: Optional[List[str]] = None) -> None:
        """
        Create a data sheet from list of dictionaries.

        Args:
            workbook: Excel workbook
            sheet_name: Name of the sheet
            data: List of dictionaries with data
            headers: Optional list of headers
        """
        ws = workbook.create_sheet(sheet_name)

        if not data:
            ws['A1'] = "No data available"
            return

        # Use provided headers or extract from first dict
        if headers is None:
            headers = list(data[0].keys()) if isinstance(data[0], dict) else []

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, record in enumerate(data, 2):
            if isinstance(record, dict):
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_num, value=record.get(header, ""))
            elif isinstance(record, (list, tuple)):
                for col_num, value in enumerate(record, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

        # Style headers
        ExcelReportGenerator.style_header(ws, row=1)
        ExcelReportGenerator.auto_size_columns(ws)

    @staticmethod
    def create_analysis_sheet(workbook, analysis: Dict[str, Any]) -> None:
        """
        Create analysis results sheet.

        Args:
            workbook: Excel workbook
            analysis: Analysis results
        """
        ws = workbook.create_sheet("Analysis")

        row = 1
        ws[f'A{row}'] = "분석 결과"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2

        # Candidate analysis
        if "candidates" in analysis:
            ws[f'A{row}'] = "후보자별 분석"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = "후보자"
            ws[f'B{row}'] = "득표수"
            ws[f'C{row}'] = "득표율 (%)"
            ExcelReportGenerator.style_header(ws, row=row)
            row += 1

            for candidate_data in analysis["candidates"]:
                ws[f'A{row}'] = candidate_data.get("name", "")
                ws[f'B{row}'] = candidate_data.get("votes", 0)
                ws[f'C{row}'] = candidate_data.get("vote_percentage", 0)
                row += 1

        # Anomalies
        if "anomalies" in analysis and analysis["anomalies"]:
            row += 2
            ws[f'A{row}'] = "이상치 탐지"
            ws[f'A{row}'].font = Font(bold=True, color="FF0000")
            row += 1

            for anomaly in analysis["anomalies"]:
                ws[f'A{row}'] = anomaly
                ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                row += 1

        ExcelReportGenerator.auto_size_columns(ws)

    @staticmethod
    def generate_report(
        output_path: str,
        raw_data: Optional[List[Dict]] = None,
        enriched_data: Optional[List[Dict]] = None,
        analysis: Optional[Dict[str, Any]] = None,
        summary: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Generate complete Excel report.

        Args:
            output_path: Path to save Excel file
            raw_data: Raw extracted data
            enriched_data: Validated and enriched data
            analysis: Analysis results
            summary: Summary data

        Returns:
            Path to generated file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        logger.info(f"Generating Excel report: {output_path}")

        wb = ExcelReportGenerator.create_workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create summary sheet
        if summary:
            ExcelReportGenerator.create_summary_sheet(wb, summary)

        # Create raw data sheet
        if raw_data:
            ExcelReportGenerator.create_data_sheet(wb, "Raw Data", raw_data)

        # Create enriched data sheet
        if enriched_data:
            ExcelReportGenerator.create_data_sheet(wb, "Enriched Data", enriched_data)

        # Create analysis sheet
        if analysis:
            ExcelReportGenerator.create_analysis_sheet(wb, analysis)

        # Save workbook
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"Excel report generated successfully: {output_path}")
        return output_path
